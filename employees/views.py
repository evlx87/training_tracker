import logging
from functools import wraps
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin, UserPassesTestMixin
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from .forms import EmployeeForm, DepartmentForm, PositionForm, TrainingProgramForm, TrainingRecordForm
from .models import Employee, Department, Position, TrainingProgram, TrainingRecord
from .services import ReportService

# Настройка логгера
logger = logging.getLogger('employees')

# Декоратор для логирования действий в представлениях
def log_view_action(action, model_name):
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(view, request, *args, **kwargs):
            user = request.user.username if request.user.is_authenticated else 'Anonymous'
            logger.info(
                '%s %s пользователем: %s, метод: %s, путь: %s, параметры: %s',
                action,
                model_name,
                user,
                request.method,
                request.path,
                request.GET.dict() if request.method == 'GET' else request.POST.dict()
            )
            try:
                response = view_func(view, request, *args, **kwargs)
                logger.debug('Успешное выполнение %s %s пользователем: %s', action, model_name, user)
                return response
            except Exception as e:
                logger.error(
                    'Ошибка при %s %s пользователем: %s, ошибка: %s',
                    action, model_name, user, str(e), exc_info=True
                )
                raise
        return wrapper
    return decorator

# Базовый класс для удаления с подтверждением от группы MTO
class MTOConfirmedDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    confirm_url_name = None
    permission_required = None

    def test_func(self):
        user = self.request.user
        has_permission = user.groups.filter(name=settings.MTO_GROUP_NAME).exists() or user.has_perm(self.permission_required)
        logger.debug(
            'Проверка прав доступа для удаления, пользователь: %s, имеет права: %s',
            user.username, has_permission
        )
        return has_permission

    def post(self, request, *args, **kwargs):
        obj = self.get_object()
        user = request.user.username
        if self.test_func():
            logger.info(
                'Удален %s: %s пользователем: %s',
                self.model._meta.verbose_name, obj, user
            )
            return super().post(request, *args, **kwargs)
        messages.error(request, 'Удаление требует подтверждения от пользователя из группы MTO.')
        logger.warning(
            'Отказано в удалении %s: %s пользователем: %s, требуется подтверждение MTO',
            self.model._meta.verbose_name, obj, user
        )
        return redirect(self.confirm_url_name, pk=obj.pk)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.path.endswith('confirm/'):
            context['confirm_mode'] = True
        return context

class IndexView(TemplateView):
    template_name = 'index.html'

    @log_view_action('Открыта', 'главная страница')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = 'employee_list.html'
    context_object_name = 'employees'
    paginate_by = 20

    def get_queryset(self):
        queryset = Employee.objects.select_related('position', 'department')
        logger.debug('Получен запрос списка сотрудников, размер: %s', queryset.count())
        return queryset

    @log_view_action('Запрошен список', 'сотрудников')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

class EmployeeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'employee_form.html'
    success_url = reverse_lazy('employees:employee_list')
    permission_required = 'employees.add_employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Добавить'
        return context

    @log_view_action('Открыта форма создания', 'сотрудника')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        employee = form.save()
        logger.info('Создан сотрудник: %s пользователем: %s', employee, self.request.user.username)
        return super().form_valid(form)

    def form_invalid(self, form):
        logger.warning(
            'Ошибка валидации формы создания сотрудника: %s пользователем: %s',
            form.errors, self.request.user.username
        )
        return super().form_invalid(form)

class EmployeeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'employee_form.html'
    success_url = reverse_lazy('employees:employee_list')
    permission_required = 'employees.change_employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Редактировать'
        return context

    @log_view_action('Открыта форма редактирования', 'сотрудника')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        employee = form.save()
        logger.info('Обновлен сотрудник: %s пользователем: %s', employee, self.request.user.username)
        return super().form_valid(form)

    def form_invalid(self, form):
        logger.warning(
            'Ошибка валидации формы редактирования сотрудника: %s пользователем: %s',
            form.errors, self.request.user.username
        )
        return super().form_invalid(form)

class EmployeeDeleteView(MTOConfirmedDeleteView):
    model = Employee
    template_name = 'employee_confirm_delete.html'
    success_url = reverse_lazy('employees:employee_list')
    confirm_url_name = 'employees:employee_delete_confirm'
    permission_required = 'employees.delete_employee'

class EmployeeDeleteConfirmView(MTOConfirmedDeleteView):
    model = Employee
    template_name = 'employee_confirm_delete.html'
    success_url = reverse_lazy('employees:employee_list')
    confirm_url_name = 'employees:employee_delete_confirm'
    permission_required = 'employees.delete_employee'

    def post(self, request, *args, **kwargs):
        obj = self.get_object()
        user = request.user.username
        if request.user.groups.filter(name=settings.MTO_GROUP_NAME).exists():
            logger.info(
                'Подтверждено удаление сотрудника: %s пользователем из группы MTO: %s', obj, user
            )
            return super(MTOConfirmedDeleteView, self).post(request, *args, **kwargs)
        messages.error(request, 'Только пользователь из группы MTO может подтвердить удаление.')
        logger.warning(
            'Отказано в подтверждении удаления сотрудника: %s пользователем: %s', obj, user
        )
        return self.render_to_response(self.get_context_data())

class EmployeeTrainingsView(LoginRequiredMixin, TemplateView):
    template_name = 'employee_trainings.html'

    def get_context_data(self, **kwargs):
        logger.debug(
            "Вызван get_context_data в EmployeeTrainingsView для pk=%s пользователем: %s",
            self.kwargs['pk'], self.request.user.username
        )
        context = super().get_context_data(**kwargs)
        employee = get_object_or_404(Employee, pk=self.kwargs['pk'])
        context['employee'] = employee
        trainings = employee.trainingrecord_set.all()
        context['training_records'] = trainings
        logger.debug(
            "Найдено %d записей об обучении для сотрудника %s", trainings.count(), employee
        )
        return context

    @log_view_action('Запрошены записи об обучении для', 'сотрудника')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

class TrainingRecordCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = TrainingRecord
    form_class = TrainingRecordForm
    template_name = 'training_record_form.html'
    permission_required = 'employees.add_trainingrecord'

    def get_employee(self):
        employee_pk = self.kwargs.get('employee_pk') or self.request.POST.get('employee_pk')
        if employee_pk:
            return get_object_or_404(Employee, pk=employee_pk)
        return None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee = self.get_employee()
        if employee:
            context['employee'] = employee
        context['action'] = 'Добавить'
        return context

    @log_view_action('Открыта форма создания записи об', 'обучении')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        employee = self.get_employee()
        user = self.request.user.username
        if not employee:
            logger.error(
                'Не указан сотрудник для создания записи об обучении пользователем: %s', user
            )
            form.add_error(None, 'Сотрудник не выбран.')
            return self.form_invalid(form)
        form.instance.employee = employee
        training_record = form.save()
        logger.info(
            'Создана запись об обучении для %s пользователем: %s', training_record.employee, user
        )
        return redirect('employees:employee_trainings', pk=training_record.employee.pk)

    def form_invalid(self, form):
        logger.warning(
            'Ошибка валидации формы создания записи об обучении: %s пользователем: %s',
            form.errors, self.request.user.username
        )
        return super().form_invalid(form)

class TrainingRecordUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = TrainingRecord
    form_class = TrainingRecordForm
    template_name = 'training_record_form.html'
    permission_required = 'employees.change_trainingrecord'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['employee'] = self.get_object().employee
        context['action'] = 'Редактировать'
        return context

    @log_view_action('Открыта форма редактирования записи об', 'обучении')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        training_record = form.save()
        logger.info(
            'Обновлена запись об обучении для %s пользователем: %s',
            training_record.employee, self.request.user.username
        )
        return redirect('employees:employee_trainings', pk=training_record.employee.pk)

    def form_invalid(self, form):
        logger.warning(
            'Ошибка валидации формы редактирования записи об обучении: %s пользователем: %s',
            form.errors, self.request.user.username
        )
        return super().form_invalid(form)

class TrainingRecordDeleteView(MTOConfirmedDeleteView):
    model = TrainingRecord
    template_name = 'training_record_confirm_delete.html'
    confirm_url_name = 'employees:training_record_delete_confirm'
    permission_required = 'employees.delete_trainingrecord'

    def get_success_url(self):
        return reverse_lazy('employees:employee_trainings', kwargs={'pk': self.get_object().employee.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['employee'] = self.get_object().employee
        return context

class TrainingRecordDeleteConfirmView(MTOConfirmedDeleteView):
    model = TrainingRecord
    template_name = 'training_record_confirm_delete.html'
    confirm_url_name = 'employees:training_record_delete_confirm'
    permission_required = 'employees.delete_trainingrecord'

    def get_success_url(self):
        return reverse_lazy('employees:employee_trainings', kwargs={'pk': self.get_object().employee.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['employee'] = self.get_object().employee
        return context

    def post(self, request, *args, **kwargs):
        obj = self.get_object()
        user = request.user.username
        if request.user.groups.filter(name=settings.MTO_GROUP_NAME).exists():
            logger.info(
                'Подтверждено удаление записи об обучении: %s пользователем из группы MTO: %s', obj, user
            )
            return super(MTOConfirmedDeleteView, self).post(request, *args, **kwargs)
        messages.error(request, 'Только пользователь из группы MTO может подтвердить удаление.')
        logger.warning(
            'Отказано в подтверждении удаления записи об обучении: %s пользователем: %s', obj, user
        )
        return self.render_to_response(self.get_context_data())

class DepartmentListView(LoginRequiredMixin, ListView):
    model = Department
    template_name = 'departments.html'
    context_object_name = 'departments'
    paginate_by = 20

    @log_view_action('Запрошен список', 'подразделений')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

class DepartmentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Department
    form_class = DepartmentForm
    template_name = 'department_form.html'
    success_url = reverse_lazy('employees:department_list')
    permission_required = 'employees.add_department'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Добавить'
        return context

    @log_view_action('Открыта форма создания', 'подразделения')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        department = form.save()
        logger.info('Создано подразделение: %s пользователем: %s', department, self.request.user.username)
        return super().form_valid(form)

    def form_invalid(self, form):
        logger.warning(
            'Ошибка валидации формы создания подразделения: %s пользователем: %s',
            form.errors, self.request.user.username
        )
        return super().form_invalid(form)

class DepartmentUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Department
    form_class = DepartmentForm
    template_name = 'department_form.html'
    success_url = reverse_lazy('employees:department_list')
    permission_required = 'employees.change_department'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Редактировать'
        return context

    @log_view_action('Открыта форма редактирования', 'подразделения')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        department = form.save()
        logger.info('Обновлено подразделение: %s пользователем: %s', department, self.request.user.username)
        return super().form_valid(form)

    def form_invalid(self, form):
        logger.warning(
            'Ошибка валидации формы редактирования подразделения: %s пользователем: %s',
            form.errors, self.request.user.username
        )
        return super().form_invalid(form)

class DepartmentDeleteView(MTOConfirmedDeleteView):
    model = Department
    template_name = 'department_confirm_delete.html'
    success_url = reverse_lazy('employees:department_list')
    confirm_url_name = 'employees:department_delete_confirm'
    permission_required = 'employees.delete_department'

class DepartmentDeleteConfirmView(MTOConfirmedDeleteView):
    model = Department
    template_name = 'department_confirm_delete.html'
    success_url = reverse_lazy('employees:department_list')
    confirm_url_name = 'employees:department_delete_confirm'
    permission_required = 'employees.delete_department'

    def post(self, request, *args, **kwargs):
        obj = self.get_object()
        user = request.user.username
        if request.user.groups.filter(name=settings.MTO_GROUP_NAME).exists():
            logger.info(
                'Подтверждено удаление подразделения: %s пользователем из группы MTO: %s', obj, user
            )
            return super(MTOConfirmedDeleteView, self).post(request, *args, **kwargs)
        messages.error(request, 'Только пользователь из группы MTO может подтвердить удаление.')
        logger.warning(
            'Отказано в подтверждении удаления подразделения: %s пользователем: %s', obj, user
        )
        return self.render_to_response(self.get_context_data())

class PositionListView(LoginRequiredMixin, ListView):
    model = Position
    template_name = 'positions.html'
    context_object_name = 'positions'
    paginate_by = 20

    @log_view_action('Запрошен список', 'должностей')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

class PositionCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Position
    form_class = PositionForm
    template_name = 'position_form.html'
    success_url = reverse_lazy('employees:position_list')
    permission_required = 'employees.add_position'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Добавить'
        return context

    @log_view_action('Открыта форма создания', 'должности')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        position = form.save()
        logger.info('Создана должность: %s пользователем: %s', position, self.request.user.username)
        return super().form_valid(form)

    def form_invalid(self, form):
        logger.warning(
            'Ошибка валидации формы создания должности: %s пользователем: %s',
            form.errors, self.request.user.username
        )
        return super().form_invalid(form)

class PositionUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Position
    form_class = PositionForm
    template_name = 'position_form.html'
    success_url = reverse_lazy('employees:position_list')
    permission_required = 'employees.change_position'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Редактировать'
        return context

    @log_view_action('Открыта форма редактирования', 'должности')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        position = form.save()
        logger.info('Обновлена должность: %s пользователем: %s', position, self.request.user.username)
        return super().form_valid(form)

    def form_invalid(self, form):
        logger.warning(
            'Ошибка валидации формы редактирования должности: %s пользователем: %s',
            form.errors, self.request.user.username
        )
        return super().form_invalid(form)

class PositionDeleteView(MTOConfirmedDeleteView):
    model = Position
    template_name = 'position_confirm_delete.html'
    success_url = reverse_lazy('employees:position_list')
    confirm_url_name = 'employees:position_delete_confirm'
    permission_required = 'employees.delete_position'

class PositionDeleteConfirmView(MTOConfirmedDeleteView):
    model = Position
    template_name = 'position_confirm_delete.html'
    success_url = reverse_lazy('employees:position_list')
    confirm_url_name = 'employees:position_delete_confirm'
    permission_required = 'employees.delete_position'

    def post(self, request, *args, **kwargs):
        obj = self.get_object()
        user = request.user.username
        if request.user.groups.filter(name=settings.MTO_GROUP_NAME).exists():
            logger.info(
                'Подтверждено удаление должности: %s пользователем из группы MTO: %s', obj, user
            )
            return super(MTOConfirmedDeleteView, self).post(request, *args, **kwargs)
        messages.error(request, 'Только пользователь из группы MTO может подтвердить удаление.')
        logger.warning(
            'Отказано в подтверждении удаления должности: %s пользователем: %s', obj, user
        )
        return self.render_to_response(self.get_context_data())

class TrainingListView(LoginRequiredMixin, ListView):
    model = TrainingProgram
    template_name = 'trainings.html'
    context_object_name = 'trainings'
    paginate_by = 20

    @log_view_action('Запрошен список', 'программ обучения')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

class TrainingCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = TrainingProgram
    form_class = TrainingProgramForm
    template_name = 'training_form.html'
    success_url = reverse_lazy('employees:training_list')
    permission_required = 'employees.add_trainingprogram'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Добавить'
        return context

    @log_view_action('Открыта форма создания', 'программы обучения')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        training = form.save()
        logger.info('Создана программа обучения: %s пользователем: %s', training, self.request.user.username)
        return super().form_valid(form)

    def form_invalid(self, form):
        logger.warning(
            'Ошибка валидации формы создания программы обучения: %s пользователем: %s',
            form.errors, self.request.user.username
        )
        return super().form_invalid(form)

class TrainingUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = TrainingProgram
    form_class = TrainingProgramForm
    template_name = 'training_form.html'
    success_url = reverse_lazy('employees:training_list')
    permission_required = 'employees.change_trainingprogram'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Редактировать'
        return context

    @log_view_action('Открыта форма редактирования', 'программы обучения')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        training = form.save()
        logger.info('Обновлена программа обучения: %s пользователем: %s', training, self.request.user.username)
        return super().form_valid(form)

    def form_invalid(self, form):
        logger.warning(
            'Ошибка валидации формы редактирования программы обучения: %s пользователем: %s',
            form.errors, self.request.user.username
        )
        return super().form_invalid(form)

class TrainingDeleteView(MTOConfirmedDeleteView):
    model = TrainingProgram
    template_name = 'training_confirm_delete.html'
    success_url = reverse_lazy('employees:training_list')
    confirm_url_name = 'employees:training_delete_confirm'
    permission_required = 'employees.delete_trainingprogram'
    context_object_name = 'training'

class TrainingDeleteConfirmView(MTOConfirmedDeleteView):
    model = TrainingProgram
    template_name = 'training_confirm_delete.html'
    success_url = reverse_lazy('employees:training_list')
    confirm_url_name = 'employees:training_delete_confirm'
    permission_required = 'employees.delete_trainingprogram'
    context_object_name = 'training'

    def post(self, request, *args, **kwargs):
        obj = self.get_object()
        user = request.user.username
        if request.user.groups.filter(name=settings.MTO_GROUP_NAME).exists():
            logger.info(
                'Подтверждено удаление программы обучения: %s пользователем из группы MTO: %s', obj, user
            )
            return super(MTOConfirmedDeleteView, self).post(request, *args, **kwargs)
        messages.error(request, 'Только пользователь из группы MTO может подтвердить удаление.')
        logger.warning(
            'Отказано в подтверждении удаления программы обучения: %s пользователем: %s', obj, user
        )
        return self.render_to_response(self.get_context_data())

class ReportsView(TemplateView):
    template_name = 'reports.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_employees = self.request.GET.getlist('employees')
        selected_program = self.request.GET.get('program')
        selected_employees = [emp for emp in selected_employees if emp]
        logger.debug("Selected employees after filtering: %s", selected_employees)
        report_data, training_programs = ReportService.generate_training_report(
            selected_employees, selected_program
        )
        if selected_employees:
            report_data = [
                data for data in report_data if str(data['employee'].pk) in selected_employees
            ]
        logger.debug("Report data length after filtering: %s", len(report_data))
        sort_by = self.request.GET.get('sort_by')
        sort_order = self.request.GET.get('sort_order', 'asc')
        if sort_by and sort_by.isdigit():
            report_data.sort(
                key=lambda x: x['trainings'].get(int(sort_by), {}).get('date', 'Обучение не пройдено'),
                reverse=(sort_order == 'desc')
            )
        context['report_data'] = report_data
        context['training_programs'] = training_programs
        context['employees'] = Employee.objects.all()
        context['departments'] = Department.objects.all()
        context['selected_employees'] = selected_employees
        context['selected_program'] = selected_program
        if selected_program and selected_program.isdigit():
            program = TrainingProgram.objects.filter(id=int(selected_program)).first()
            context['selected_program_name'] = program.name if program else "Неизвестная программа"
        total_employees = len(report_data)
        context['total_employees'] = total_employees
        trained_counts = {}
        for program in training_programs:
            trained_count = sum(
                1 for data in report_data if data['trainings'].get(program.id, {}).get('date') != "Обучение не пройдено"
            )
            trained_counts[program.name] = trained_count
        context['trained_counts'] = trained_counts
        return context

    @log_view_action('Открыта страница', 'отчетов')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

class ExportReportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        report_data, training_programs = ReportService.generate_training_report()
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет по обучению"
        headers = ["Сотрудник", "Должность", "Подразделение"] + [program.name for program in training_programs]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for data in report_data:
            first_initial = data['employee'].first_name[0] if data['employee'].first_name else ""
            middle_initial = data['employee'].middle_name[0] if data['employee'].middle_name else ""
            employee_name = f"{data['employee'].last_name} {first_initial}. {middle_initial}.".strip()
            row = [
                employee_name,
                str(data['employee'].position or "—"),
                str(data['employee'].department or "—")
            ]
            for program in training_programs:
                training = data['trainings'].get(program.id, {})
                date = training.get('date', "Обучение не пройдено")
                row.append(date if date == "Обучение не пройдено" else date.strftime("%d.%m.%y"))
            ws.append(row)
        for row_idx, data in enumerate(report_data, start=2):
            for col_idx, program in enumerate(training_programs, start=4):
                cell = ws.cell(row=row_idx, column=col_idx)
                training = data['trainings'].get(program.id, {})
                status_class = training.get('class', 'not-completed')
                fill_colors = {
                    'not-completed': 'FF9999',
                    'overdue': 'FF3333',
                    'warning': 'FFFF66',
                    'completed': '99FF99'
                }
                cell.fill = PatternFill(
                    start_color=fill_colors.get(status_class, 'FFFFFF'),
                    end_color=fill_colors.get(status_class, 'FFFFFF'),
                    fill_type="solid"
                )
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="training_report.xlsx"'
        wb.save(response)
        logger.info('Экспортирован отчет по обучению пользователем: %s', request.user.username)
        return response

    @log_view_action('Экспортирован', 'отчет по обучению')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)