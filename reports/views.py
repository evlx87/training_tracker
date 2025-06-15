import logging

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.views import View
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from departments.models import Department
from employees.models import Employee
from reports.services import ReportService
from employees.views import log_view_action
from trainings.models import TrainingProgram

logger = logging.getLogger('reports')


class ReportsView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/report_list.html'
    permission_required = 'employees.view_report'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_employees = self.request.GET.getlist('employees')
        selected_program = self.request.GET.get('program')
        exclude_not_completed = self.request.GET.get(
            'exclude_not_completed') == 'on'
        selected_employees = [emp for emp in selected_employees if emp]
        logger.debug(
            "Selected employees after filtering: %s",
            selected_employees)
        report_data, training_programs = ReportService.generate_training_report(
            selected_employees, selected_program)

        if selected_employees:
            report_data = [
                data for data in report_data if str(
                    data['employee'].pk) in selected_employees]

        if exclude_not_completed:
            if selected_program and selected_program.isdigit():
                # Исключаем сотрудников, у которых нет записи по выбранной
                # программе
                report_data = [
                    data for data in report_data
                    if data['trainings'].get(int(selected_program), {}).get('date') != "Обучение не пройдено"
                ]
            else:
                # Исключаем сотрудников, у которых нет ни одной записи об
                # обучении
                report_data = [
                    data for data in report_data
                    if any(
                        training.get('date') != "Обучение не пройдено"
                        for training in data['trainings'].values()
                    )
                ]
        logger.debug(
            "Report data length after filtering: %s",
            len(report_data))
        sort_by = self.request.GET.get('sort_by')
        sort_order = self.request.GET.get('sort_order', 'asc')

        if sort_by and sort_by.isdigit() and int(
                sort_by) in [program.id for program in training_programs]:
            report_data.sort(
                key=lambda x: x['trainings'].get(
                    int(sort_by), {}).get(
                    'date', 'Обучение не пройдено'),
                reverse=(sort_order == 'desc')
            )
        context['report_data'] = report_data
        context['training_programs'] = training_programs
        context['employees'] = Employee.objects.all()
        context['departments'] = Department.objects.all()
        context['selected_employees'] = selected_employees
        context['selected_program'] = selected_program
        context['exclude_not_completed'] = exclude_not_completed

        if selected_program and selected_program.isdigit():
            program = TrainingProgram.objects.filter(
                id=int(selected_program)).first()
            context['selected_program_name'] = program.name if program else "Неизвестная программа"

        total_employees = len(report_data)
        context['total_employees'] = total_employees
        trained_counts = {}
        for program in training_programs:
            trained_count = sum(
                1 for data in report_data if data['trainings'].get(
                    program.id, {}).get('date') != "Обучение не пройдено")
            trained_counts[program.name] = trained_count
        context['trained_counts'] = trained_counts
        return context

    @log_view_action('Открыта страница', 'отчетов')
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


class ExportReportView(LoginRequiredMixin, View):
    @log_view_action('Экспортирован', 'отчет по обучению')
    def get(self, request, *args, **kwargs):
        try:
            selected_employees = request.GET.getlist('employees')
            selected_program = request.GET.get('program')
            exclude_not_completed = request.GET.get('exclude_not_completed') == 'on'

            report_data, training_programs = ReportService.generate_training_report(
                selected_employees, selected_program)

            # Фильтрация данных по сотрудникам
            if selected_employees:
                report_data = [
                    data for data in report_data if str(data['employee'].pk) in selected_employees]

            # Фильтрация по exclude_not_completed
            if exclude_not_completed:
                if selected_program and selected_program.isdigit():
                    report_data = [
                        data for data in report_data
                        if data['trainings'].get(int(selected_program), {}).get('date') != "Обучение не пройдено"
                    ]
                else:
                    report_data = [
                        data for data in report_data
                        if any(
                            training.get('date') != "Обучение не пройдено"
                            for training in data['trainings'].values()
                        )
                    ]

            wb = Workbook()
            ws = wb.active
            ws.title = "Отчет по обучению"

            # Определяем заголовки
            base_headers = ["Сотрудник", "Должность", "Руководитель", "Педагогический работник",
                            "Член комиссии по ОТ", "Подразделение"]
            if selected_program and selected_program.isdigit():
                # Если выбрана программа, добавляем только ее
                program = TrainingProgram.objects.filter(id=int(selected_program)).first()
                program_headers = [program.name] if program else []
            else:
                # Иначе добавляем все программы
                program_headers = [program.name for program in training_programs]
            headers = base_headers + program_headers
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Формируем строки данных
            for data in report_data:
                first_initial = data['employee'].first_name[0] if data['employee'].first_name else ""
                middle_initial = data['employee'].middle_name[0] if data['employee'].middle_name else ""
                employee_name = f"{data['employee'].last_name} {first_initial}. {middle_initial}.".strip()
                row = [
                    employee_name,
                    str(data['employee'].position or "—"),
                    "Да" if data['employee'].position and data['employee'].position.is_manager else "Нет",
                    "Да" if data['employee'].position and data['employee'].position.is_teacher else "Нет",
                    "Да" if data['employee'].is_safety_commission_member else "Нет",
                    str(data['employee'].department or "—")
                ]
                # Добавляем данные по программам
                if selected_program and selected_program.isdigit():
                    # Только для выбранной программы
                    training = data['trainings'].get(int(selected_program), {})
                    date = training.get('date', "Обучение не пройдено")
                    row.append(date if date == "Обучение не пройдено" else date.strftime("%d.%m.%y"))
                else:
                    # Для всех программ
                    for program in training_programs:
                        training = data['trainings'].get(program.id, {})
                        date = training.get('date', "Обучение не пройдено")
                        row.append(date if date == "Обучение не пройдено" else date.strftime("%d.%m.%y"))
                ws.append(row)

            # Применяем цветовую заливку для колонок с программами
            start_col_idx = len(base_headers) + 1  # Индекс первой колонки с программой
            for row_idx, data in enumerate(report_data, start=2):
                if selected_program and selected_program.isdigit():
                    # Только для выбранной программы
                    cell = ws.cell(row=row_idx, column=start_col_idx)
                    training = data['trainings'].get(int(selected_program), {})
                    status_class = training.get('class', 'not-completed')
                    fill_colors = {
                        'not-completed': 'FF9999',
                        'overdue': 'FF3333',
                        'warning': 'FFFF66',
                        'completed': '99FF99'
                    }
                    cell.fill = PatternFill(
                        start_color=fill_colors.get(status_class, 'FFFFFF'),
                        end_color=fill_colors.get(status_class, 'FFFFFF'),
                        fill_type="solid"
                    )
                else:
                    # Для всех программ
                    for col_idx, program in enumerate(training_programs, start=start_col_idx):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        training = data['trainings'].get(program.id, {})
                        status_class = training.get('class', 'not-completed')
                        fill_colors = {
                            'not-completed': 'FF9999',
                            'overdue': 'FF3333',
                            'warning': 'FFFF66',
                            'completed': '99FF99'
                        }
                        cell.fill = PatternFill(
                            start_color=fill_colors.get(status_class, 'FFFFFF'),
                            end_color=fill_colors.get(status_class, 'FFFFFF'),
                            fill_type="solid"
                        )

            # Настраиваем ширину колонок
            for col in ws.columns:
                max_length = max(len(str(cell.value)) for cell in col if cell.value)
                ws.column_dimensions[col[0].column_letter].width = max_length + 2

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            # Динамическое имя файла
            filename = "training_report_filtered.xlsx" if selected_employees or selected_program or exclude_not_completed else "training_report_all.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            logger.info('Экспортирован %s отчет по обучению пользователем: %s',
                        'отфильтрованный' if selected_employees or selected_program or exclude_not_completed else 'полный',
                        request.user.username)
            return response
        except Exception as e:
            logger.error(f"Ошибка при экспорте отчета: {e}")
            return HttpResponse("Ошибка при создании отчета. Пожалуйста, попробуйте позже.", status=500)